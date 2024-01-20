from forms2 import *
from datetime import datetime
import pandas as pd
import openpyxl as px
import requests
from requests.auth import HTTPBasicAuth
import credentials
from difflib import SequenceMatcher
import mappings
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# nc-credentials
NEXTCLOUD_USERNAME = credentials.username
NEXTCLOUD_PASSWORD = credentials.password
NEXTCLOUD_URL = credentials.url4
# WebDAV URL
webdav_url = credentials.url_elternaccounts
webdav_url2 = credentials.url_elternaccounts2

# Zugangsdaten
# username = credentials.elternzugaenge_username
# password = credentials.elternzugaenge_pw


def get_file(url, filename, username, password):
    # Anfrage, um die Datei zu bekommen
    response = requests.get(url, auth=HTTPBasicAuth(username, password))

    if response.status_code == 200:
        # Erfolgreich die Datei erhalten
        file_content = response.content
        # Schreiben des Inhalts in eine neue Datei
        with open(filename, "wb") as file:
            file.write(file_content)
        print(f"Datei wurde heruntergeladen von {url}")
    else:
        print(f"Fehler beim Zugriff auf die Datei: {response.status_code}")


def put_file(url, filename, username, password):
    with open(filename, "rb") as file:
        file_content = file.read()
    # Anfrage, um die Datei zu schicken
    response = requests.put(
        url, data=file_content, auth=HTTPBasicAuth(username, password)
    )

    if (
        response.status_code == 200
        or response.status_code == 201
        or response.status_code == 204
    ):
        # Erfolgreich die Datei gesendet
        print(f"Datei wurde hochgeladen zu {url}")
    else:
        print(f"Fehler beim Zugriff auf die Datei: {response.status_code}")


def update_xlsx(csv_path, xlsx_path):
    # Aktuelles Datum und Zeit holen
    now = datetime.now()

    # Datum und Zeit im Format yymmddhhmmss formatieren
    date_string = now.strftime("%y%m%d%H%M%S")

    # Die CSV-Datei lesen
    csv_df = pd.read_csv(csv_path)

    # Die XLSX-Datei lesen
    xlsx_df = pd.read_excel(xlsx_path)
    # TODO: Hier den Speicherort anpassen mit webdav
    # print(webdav_url)
    put_file(
        f"{webdav_url[:-5]}_backup{date_string}.xlsx",
        "testxlsx.xlsx",
        NEXTCLOUD_USERNAME,
        NEXTCLOUD_PASSWORD,
    )
    # xlsx_df.to_excel(f"{xlsx_path[:-5]}_backup{date_string}.xlsx", index=False)

    # Hier fügen wir die Spalte "Kontrolliert" zur CSV-Datei hinzu.
    csv_df["Kontrolliert"] = pd.NA

    # Die beiden DataFrames zusammenführen, wobei die ursprünglichen Werte in der XLSX-Datei beibehalten werden
    merged_df = pd.concat([xlsx_df, csv_df], ignore_index=True)

    # Duplikate entfernen, wobei die erste Instanz eines Duplikats beibehalten wird
    merged_df.drop_duplicates(subset="Zeitstempel", keep="first", inplace=True)

    # Die aktualisierte XLSX-Datei schreiben
    merged_df.to_excel(xlsx_path, index=False)

    # Öffnen Sie die Arbeitsmappe und wählen Sie ein Arbeitsblatt aus
    wb = px.load_workbook(xlsx_path)
    ws = wb.active

    # Erstellen Sie einen AutoFilter für alle Spalten in Ihrem Bereich
    ws.auto_filter.ref = ws.dimensions

    # AutoFilter erstellen
    ws.auto_filter.ref = ws.dimensions

    # Spaltenbreiten für "Emailadresse des Elternteil" und "alle Klassen der Kinder" anpassen
    for column in [
        "F",
        "P",
        "I",
        "L",
        "O",
    ]:  # Angenommen, diese Spalten enthalten die relevanten Daten
        max_length = 0
        column_cells = ws[column]
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # # Spaltenbreiten anpassen
    # for column in ws.columns:
    #     max_length = 0
    #     column = [cell for cell in column]
    #     for cell in column:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(cell.value)
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2)
    #     ws.column_dimensions[px.utils.get_column_letter(column[0].column)].width = adjusted_width

    # Spaltenbreiten für "Benutzer-ID", "Anzeigename des Nutzers" und "Zeitstempel" auf 1 cm (etwa 3 Zeichen) setzen
    ws.column_dimensions[px.utils.get_column_letter(ws["A"][0].column)].width = 2
    ws.column_dimensions[px.utils.get_column_letter(ws["B"][0].column)].width = 2
    ws.column_dimensions[px.utils.get_column_letter(ws["C"][0].column)].width = 2
    ws.column_dimensions[px.utils.get_column_letter(ws["D"][0].column)].width = 9
    ws.column_dimensions[px.utils.get_column_letter(ws["E"][0].column)].width = 9
    ws.column_dimensions[px.utils.get_column_letter(ws["G"][0].column)].width = 15
    ws.column_dimensions[px.utils.get_column_letter(ws["H"][0].column)].width = 9
    ws.column_dimensions[px.utils.get_column_letter(ws["J"][0].column)].width = 9
    ws.column_dimensions[px.utils.get_column_letter(ws["K"][0].column)].width = 9
    ws.column_dimensions[px.utils.get_column_letter(ws["M"][0].column)].width = 9
    ws.column_dimensions[px.utils.get_column_letter(ws["N"][0].column)].width = 9
    # ws.column_dimensions[px.utils.get_column_letter(ws["I"][0].column)].width = 11
    # ws.column_dimensions[px.utils.get_column_letter(ws["P"][0].column)].width = 11
    ws.column_dimensions[px.utils.get_column_letter(ws["Q"][0].column)].width = 11

    # Speichern Sie die Arbeitsmappe
    wb.save(xlsx_path)


def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()


def returnUsername(given, last, typ):
    if typ == "vorname.nachname":
        username = (
            given.translate(mappings.mappingusername).split(" ")[0]
            + "."
            + last.translate(mappings.mappingusername).replace(" ", "").replace("-", "")
        )
        # print(f"{username}")
        if username in mappings.alternative_usernames:
            username = mappings.alt_usernames[username]
            # print(f"Habe alternativen username gespeichert: {username}")
    if typ == "kurzform":
        username = (
            given.translate(mappings.mappingusername).split(" ")[0][:4]
            + last.translate(mappings.mappingusername)
            .replace(" ", "")
            .replace("-", "")[:4]
        )
    return username.lower()


def createElternaccounts(formsdatei, schildexport, kontrolloutput, outputfile):
    # Einlesen der XLSX-Datei
    forms = pd.read_excel(formsdatei)
    schild = pd.read_csv(schildexport, delimiter=";", quotechar='"')

    outputtest = []
    output = []

    # Filtern nach der Spalte "Kontrolliert"
    forms_filtered = forms[forms["Kontrolliert"] == 1]

    for idx, form_row in forms_filtered.iterrows():
        for i in range(1, 4):
            fname = form_row[f"Vorname des {i}. Kindes"]
            lname = form_row[f"Nachname des {i}. Kindes"]
            klasse = form_row[f"Klasse des {i}. Kindes"]

            matching_schild = None
            second_matching_schild = None
            highest_similarity = 0
            second_highest_similarity = 0
            for _, schild_row in schild.iterrows():
                if schild_row["webuntisKlasse"] == klasse:
                    full_name_schild = (
                        f"{schild_row['US_firstName']} {schild_row['US_lastName']}"
                    )
                    full_name_forms = f"{fname} {lname}"
                    similarity = similar(full_name_schild, full_name_forms)
                    if similarity > highest_similarity:
                        second_highest_similarity = highest_similarity
                        highest_similarity = similarity
                        second_matching_schild = matching_schild
                        matching_schild = schild_row
                    elif similarity > second_highest_similarity:
                        second_highest_similarity = similarity
                        second_matching_schild = schild_row

            if matching_schild is not None:
                if highest_similarity < 0.9:
                    print(
                        form_row["Vorname des Elternteils"],
                        form_row["Nachname des Elternteils"],
                        fname,
                        lname,
                        matching_schild["US_firstName"],
                        matching_schild["US_lastName"],
                        matching_schild["AT_webuntisUid"],
                        highest_similarity,
                        second_highest_similarity,
                    )
                outputtest.append(
                    [
                        form_row["Vorname des Elternteils"],
                        form_row["Nachname des Elternteils"],
                        fname,
                        lname,
                        matching_schild["US_firstName"],
                        matching_schild["US_lastName"],
                        matching_schild["AT_webuntisUid"],
                        highest_similarity,
                        second_highest_similarity,
                    ]
                )
                output.append(
                    [
                        form_row["Vorname des Elternteils"],
                        form_row["Nachname des Elternteils"],
                        form_row["Emailadresse des Elternteils"].lower(),
                        matching_schild["AT_webuntisUid"],
                    ]
                )

    output_df = pd.DataFrame(
        outputtest,
        columns=[
            "Eltern Vorname",
            "Eltern Nachname",
            "Kind Vorname (forms)",
            "Kind Nachname (forms)",
            "Kind Vorname (schild)",
            "Kind Nachname (schild)",
            "AT_webuntisUid",
            "Best Similarity Score",
            "Second Best Similarity Score",
        ],
    )
    output_df.to_csv(kontrolloutput, index=False, sep=";")
    output_df2 = pd.DataFrame(
        output, columns=["Eltern Vorname", "Eltern Nachname", "email", "student-id"]
    )
    output_df2["username"] = output_df2.apply(
        lambda row: returnUsername(
            row["Eltern Vorname"], row["Eltern Nachname"], "kurzform"
        ),
        axis=1,
    )
    # print(output_df2)
    output_df2.to_csv(outputfile, index=False, sep=";")


import re


def finde_email_adressen(text):
    # RegEx-Muster für eine einfache E-Mail-Adresse
    muster = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"

    # Verwende findall, um alle Instanzen zu finden, die dem Muster entsprechen
    email_adressen = re.findall(muster, text)

    return email_adressen


def sende_email(empfaenger_liste, betreff, nachricht, smtp_server, smtp_port, benutzername, passwort):
    if not empfaenger_liste:
        return "Keine Empfängeradresse vorhanden."

    empfaenger = empfaenger_liste[0]
    bcc = empfaenger_liste[1:]

    msg = MIMEMultipart()
    msg['From'] = benutzername
    msg['To'] = empfaenger
    msg['Bcc'] = ", ".join(bcc)
    msg['Subject'] = betreff

    msg.attach(MIMEText(nachricht, 'plain'))

    try:
        # Verbindung zum SMTP-Server mit SSL
        server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        server.login(benutzername, passwort)

        # E-Mail senden
        server.send_message(msg)
        server.quit()
        return "E-Mail erfolgreich gesendet."
    except Exception as e:
        return f"Fehler beim Senden der E-Mail: {e}"


if __name__ == "__main__":
    # ncapi = NextcloudFormsAPI(NEXTCLOUD_URL, NEXTCLOUD_USERNAME, NEXTCLOUD_PASSWORD)
    # file_content = ncapi.getFormSubmissionsCSV(credentials.elternaccounts).content
    # with open("testforms.csv", "wb") as file:
    #     file.write(file_content)
    # print("Datei wurde gespeichert")
    # xlsxfile = get_file(
    #     webdav_url2, "testxlsx.xlsx", NEXTCLOUD_USERNAME, NEXTCLOUD_PASSWORD
    # )
    # update_xlsx("testforms.csv", "testxlsx.xlsx")
    # put_file(webdav_url, "testxlsx.xlsx", NEXTCLOUD_USERNAME, NEXTCLOUD_PASSWORD)
    # put_file(webdav_url2, "testxlsx.xlsx", NEXTCLOUD_USERNAME, NEXTCLOUD_PASSWORD)

    # get_file(
    #     credentials.url_export, "Export.csv", NEXTCLOUD_USERNAME, NEXTCLOUD_PASSWORD
    # )
    # createElternaccounts(
    #     "testxlsx.xlsx",
    #     f"Export.csv",
    #     f"elternaccounts-control.csv",
    #     f"elternaccounts.csv",
    # )
    # put_file(
    #     credentials.url_elterncsv,
    #     "elternaccounts.csv",
    #     NEXTCLOUD_USERNAME,
    #     NEXTCLOUD_PASSWORD,
    # )
    # put_file(
    #     credentials.url_elterncsvcontrol,
    #     "elternaccounts-control.csv",
    #     NEXTCLOUD_USERNAME,
    #     NEXTCLOUD_PASSWORD,
    # )
    # Beispielaufruf der Funktion
    # Beachten Sie, dass Sie gültige SMTP-Serverdetails einfügen müssen
    smtp_server = credentials.smtp_server
    smtp_port = credentials.smtp_port
    benutzername = credentials.mail_benutzername
    passwort = credentials.mail_passwort

    betreff = "Webuntis-Elternaccounts wurden erstellt"
    nachricht = """
Guten Tag,

die Webuntis-Elternaccounts wurden erstellt. Bitte folgen Sie der Anleitung hier um ihren Account einzurichten.

Freundliche Grüße
Till Lieber
"""

    # Hier die vorher extrahierte E-Mail-Liste einfügen
    empfaenger_liste = [
        "dummymail@luisengym.de"
    ]

    ergebnis = sende_email(
        empfaenger_liste,
        betreff,
        nachricht,
        smtp_server,
        smtp_port,
        benutzername,
        passwort,
    )
    print(ergebnis)
